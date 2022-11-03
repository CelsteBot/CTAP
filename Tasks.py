from multiprocessing import Process, Manager
import openpyxl.styles as styles
import openpyxl as opx
import ctypes as ctp
import pandas as pd
import colorutils
import time
import json
import csv
import os


class Tasks:
    def __init__(self, tmp_time) -> None:
        self.logs = ''
        self.tmp_time = tmp_time
        self.time = time.time()
        self.percent = 0
        self.dd_percent = 0
        self.pe_percent = 0
        self.ee_percent = 0
        self.id_field = ''

        self.settings = {}
        self.data = {}
        self.prt_data = {}
        self.processes = []
        self.dups = []
        self.fields = []

        with open(f'settings.json', 'r') as in_file:
            self.settings = json.load(in_file)
        with open(f'tmp_data{self.tmp_time}.json', 'r') as in_file:
            self.data = json.load(in_file)

        self.pmap = self.settings['ADVANCED-percent_map']
        self.dd_pmap = self.settings['ADVANCED-dedup_percent_map']
        self.pe_pmap = self.settings['ADVANCED-partner_evaluation_percent_map']
        self.ee_pmap = self.settings['ADVANCED-employer_emails_percent_map']
        self.hue_range = self.settings['ADVANCED-dd_huemap']

# Helper functions ---------------------------------------------------------------------------------
    def write_data(self, dups, solid_color, op_color, sheet, row):
        if len(dups) == 0:
            return row

        header_row = row
        row += 1

        for i, dup in enumerate(dups):
            for x, key in enumerate(dup.keys()):
                sheet.cell(row=row, column=x + 1, value=dup[key])
                sheet[row][x].fill = styles.PatternFill(start_color=op_color, end_color=op_color, fill_type='solid')
            row += 1

        for cell in sheet[header_row]:
            cell.fill = styles.PatternFill(start_color=solid_color, end_color=solid_color, fill_type='solid')

        print(sheet.cell(row=4, column=2).value)

        return row

    def update_data(self):
        with open(f'tmp_data{self.tmp_time}.json', 'w') as out_file:
            out_file.write(json.dumps(self.data, indent=4))

    def log(self, msg, error = False):
        if error:
            self.data['processing_error'] = msg
            msg = f'------------------------------------------\n{msg}\n------------------------------------------'
            with open(f'tmp_data{self.tmp_time}.json', 'w') as out_file:
                out_file.write(json.dumps(self.data, indent=4))
            
        self.logs += msg + '\n'

    def map_range(self, x, a1, a2, b1, b2):
        return b1 + ((x - a1) * (b2 - b1) / (a2 - a1))

    def update_percent(self, process, task, task_p, set = False):
        # evaluate individual percents
        match process:
            case 'dd':
                self.dd_percent = task_p if set else self.map_range(task_p, 0, 100, self.dd_pmap[task-1], self.dd_pmap[task])
            case 'pe':
                pass
            case 'ee':
                pass

        # Re-evaluate full percent
        count = len(self.processes)
        processes_p = self.dd_percent * (1/count) + self.pe_percent * (1/count) + self.ee_percent * (1/count)
        self.percent = self.map_range(processes_p, 0, 100, self.pmap[1], self.pmap[2])
        self.data['processing_percent'] = self.percent
        self.update_data()

    def field_compare(self, a: str, b: str, threshold):
        length = len(a)
        if len(str(a)) < len(str(b)):
            length = len(str(b))
        n_threshold = int(length - length * threshold)

        a1 = a.encode('utf8')
        b1 = b.encode('utf8')

        return self.MyLev(a1, b1, n_threshold)

    def full_check(self, prt_start, prt_end, dups):
        print('process start')
        self.MyLev = ctp.cdll.LoadLibrary('./my_lib.so').MyLev
        self.MyLev.argtypes = [ctp.c_char_p, ctp.c_char_p, ctp.c_uint]
        if prt_end >= len(self.prt_data):
            prt_end = len(self.data) - 1
        prt_data = self.prt_data[prt_start:prt_end]
        check_data = self.prt_data[prt_start:]
        big_msg = ''
        for i, row in enumerate(prt_data):
            check_data.pop(0)

            # update percent
            if prt_start == 0 and i % 10 == 0:
                percent = 100 * i / len(prt_data)
                self.update_percent('dd', 1, percent)

            # do checks
            for x, check_row in enumerate(check_data):
                scale = 0
                if row[self.id_key] == check_row[self.id_key]:
                    scale = len(dups) - 1
                else:
                    for field in self.fields:
                        if field['match_strength'] != 100:
                            if self.field_compare(row[field['field_name']], check_row[field['field_name']], float(field['match_strength']) / 100):
                                scale += 1
                        else:
                            if row[field['field_name']] == check_row[field['field_name']]:
                                scale += 1
                
                if scale > self.data['dd_scale_threshold']:
                    big_msg += str(row[self.id_key]) + ' has scale ' + str(scale) + '\n'
                    dups[scale].append(row)
                    dups[scale].append(check_row)

        del self.MyLev
        print('process end')
        return big_msg

# Process Functions ---------------------------------------------------------------------------------
    def dd(self):

            # Data Collection ---------------------------------------------------------------------------
            dd_start = time.time()

            # Identify Duplicates ------------------------------------------------------------------------------
            # Verify field names
            self.id_key = list(self.prt_data[0].keys())[0]
            for field in self.data['dd_settings']:
                if field['field_name'] in self.prt_data[0]:
                    self.log('Found ' + field['field_name'] + ' in row')
                    self.fields.append(field)
                else:
                    self.log('Error: ' + field['field_name'] + ' not a header in sheet!', True)
                    self.update_data()

            self.data['processing_status'] = 'Identifying Duplicates'
            self.update_data()

            # dd modes
            def adj_check():
                self.MyLev = ctp.cdll.LoadLibrary('./my_lib.so').MyLev
                self.MyLev.argtypes = [ctp.c_char_p, ctp.c_char_p, ctp.c_uint]
                check_data = self.prt_data[:self.data['dd_adjacencies']]
                self.prt_data = self.prt_data[self.data['dd_adjacencies']:]
                big_msg = ''
                for i, row in enumerate(self.prt_data):
                    # update percent
                    if i % 100 == 0:
                        percent = 100 * i / len(self.prt_data)
                        self.update_percent('dd', 1, percent)

                    # do checks
                    for x, check_row in enumerate(check_data):

                        scale = 0
                        if row[self.id_key] == check_row[self.id_key]:
                            scale = len(self.dups) - 1
                        else:
                            for field in self.fields:
                                if field['match_strength'] != 100:
                                    if self.field_compare(row[field['field_name']], check_row[field['field_name']], float(field['match_strength']) / 100):
                                        scale += 1
                                else:
                                    if row[field['field_name']] == check_row[field['field_name']]:
                                        scale += 1
                        
                        if scale > self.data['dd_scale_threshold']:
                            big_msg += str(row[self.id_key]) + ' has scale ' + str(scale) + '\n'
                            self.dups[scale].append(row)
                            self.dups[scale].append(check_row)

                    check_data.pop(0)
                    check_data.append(row)
                del self.MyLev
                return big_msg

            # get data specifically for the loop
            big_msg = ''
            self.prt_data.pop(0)
            if self.data['dd_adjacencies'] > 0 and len(self.prt_data) > self.data['dd_adjacencies']:
                self.dups = []
                for i in range(len(self.data['dd_settings']) + 1):
                    self.dups.append([])
                big_msg = adj_check()
            else:
                thread_count, over = divmod(len(self.prt_data), self.settings['HIDDEN-thread_threshold'])
                if over > 0:
                    thread_count += 1

                with Manager() as manager:
                    ts = []
                    proxy_dups = manager.list()
                    for i in range(len(self.data['dd_settings']) + 1):
                        proxy_dups.append(manager.list())
                    for i in range(0, thread_count):
                        t = Process(target=self.full_check, args=(i * self.settings['HIDDEN-thread_threshold'], (i + 1) * self.settings['HIDDEN-thread_threshold'], proxy_dups))
                        ts.append(t)
                        t.start()
                    for t in ts:
                        t.join()

                    self.dups = [[] for i in range(len(proxy_dups))]
                    for i in range(len(proxy_dups)):
                        self.dups[i] = list(proxy_dups[i])

            identify_end = time.time()

            self.log(big_msg + '\n\n\n\n\n\n\n')


            # Write duplicates to sheet ------------------------------------------------------------------------------
            # Remove unlikely dups
            copy_start = time.time()
            for i in range(self.data['dd_scale_threshold']):
                self.dups.pop(i)

            # set up workbook
            row = 1
            path = os.getcwd() + '\\tmp_' + self.settings['ADVANCED-dd_tmp_file_name'] + self.tmp_time + '.xlsx'
            wb = opx.Workbook()
            ws = wb.active

            self.data['processing_status'] = 'Constructing Dups Excel Sheet'
            for i, l in enumerate(self.dups):
                # update percent
                percent = 100 * i / len(self.dups)
                self.update_percent('dd', 2, percent)

                # calculate hue
                h_range = abs(self.hue_range[0] - self.hue_range[1])
                hue = int((i + 1) / len(self.dups) * h_range)
                hue = h_range - hue

                solid_color = colorutils.hsv_to_hex((hue, 1, 1)).lstrip('#')
                op_color = colorutils.hsv_to_hex((hue, .5, 1)).lstrip('#')

                # write data
                row = self.write_data(l, solid_color, op_color, ws, row)

            wb.save(path)

            copy_end = time.time()
            self.log('Dup Identify Time: ' + str(identify_end-dd_start))
            self.log('Dup Copy Time: ' + str(copy_end-copy_start))
            self.log('Full Time: ' + str(copy_end-dd_start))

            with open(f'Logs/DeDupLog{self.tmp_time}.txt', 'w') as out_file:
                out_file.write(self.logs)

            self.update_percent('dd', 2, 100, True)


    def open_files(self):
        self.data['processing_status'] = 'Opening Sheet'
        self.update_data()

        # check if file is a csv. If not, convert
        path = self.data['prt_path']
        if path.endswith(('.xlsx', '.xls')):
            self.data['processing_status'] = 'Converting File Sheet'
            self.update_data()

            read_file = pd.read_excel(path)
            read_file.to_csv('tmp_prt_data.csv', index=None, header=True)
            path = 'tmp_prt_data.csv'

        self.data['processing_percent'] += self.dd_pmap[0]

        with open(path, 'r') as in_file:
            self.prt_data = list(csv.DictReader(in_file))

        for p in self.processes:
            t = Process(target=eval(f'self.{p}'))
            t.start()

    def main(self):
        # Find out what functions I have to do
        if 'dd_config' in self.data['flow']:
            self.processes.append('dd')
        if 'pe_config' in self.data['flow']:
            self.processes.append('pe')
        if 'ee_config' in self.data['flow']:
            self.processes.append('ee')        
        print('im the second icky baby')

        t = Process(target = self.open_files, args=())
        t.start()


#DeDup()



